from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_FILE = Path("bundle_planning_template.xlsx")
MAX_ROWS = 301


def style_header(ws, cells):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in cells:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_input_sheet(wb):
    ws = wb.active
    ws.title = "Input_Data"
    headers = [
        "clean_product_name",
        "category",
        "product_id",
        "purchased_item_count",
        "purchased_qty_m1",
        "purchased_qty_m2",
        "purchased_qty_m3",
        "purchased_qty_m4",
        "cont_from_total",
        "stock",
        "reserved_stock",
        "available_stock",
    ]
    ws.append(headers)
    style_header(ws, ws[1])

    sample_rows = [
        ["Sparkling Water 330ml", "beverages", "P-1001", 5210, 520, 560, 610, 640, 0.091, 1400, 220, 1180],
        ["Premium Coffee Beans 1kg", "beverages", "P-1002", 3930, 215, 201, 184, 175, 0.044, 520, 55, 465],
        ["Dry Fruit Mix 200g", "snacks", "P-2001", 880, 64, 58, 52, 45, 0.021, 980, 75, 905],
        ["Sea Salt Crackers", "snacks", "P-2002", 6420, 590, 615, 640, 672, 0.102, 1700, 260, 1440],
        ["Aloe Vera Gel 250ml", "personal care", "P-3001", 710, 24, 21, 18, 15, 0.013, 760, 105, 655],
    ]
    for row in sample_rows:
        ws.append(row)

    for row in range(2, MAX_ROWS):
        ws[f"I{row}"].number_format = "0.00%"

    ws.freeze_panes = "A2"
    set_widths(
        ws,
        {
            1: 32,
            2: 20,
            3: 14,
            4: 20,
            5: 16,
            6: 16,
            7: 16,
            8: 16,
            9: 16,
            10: 12,
            11: 14,
            12: 15,
        },
    )


def build_scoring_sheet(wb):
    ws = wb.create_sheet("Scoring_Model")
    headers = [
        "product_id",
        "clean_product_name",
        "category",
        "purchased_item_count",
        "sold_qty_last_4m",
        "avg_monthly_qty_4m",
        "Movement_Percentile",
        "Slow_Mover_Score",
        "cont_from_total",
        "Contribution_Percentile",
        "stock",
        "reserved_stock",
        "available_stock",
        "available_stock_calc",
        "stock_data_check",
        "stock_coverage_months",
        "Stock_Pressure_Percentile",
        "Bundle_Priority_Score",
        "Movement_Band",
        "Bundle_Value_Cluster",
        "Anchor_Eligible",
        "Candidate_Eligible",
        "Suggested_Discount_%",
        "Anchor_Key",
        "Anchor_Name",
    ]
    ws.append(headers)
    style_header(ws, ws[1])

    for row in range(2, MAX_ROWS):
        ws[f"A{row}"] = f"=Input_Data!C{row}"
        ws[f"B{row}"] = f"=Input_Data!A{row}"
        ws[f"C{row}"] = f"=Input_Data!B{row}"
        ws[f"D{row}"] = f"=Input_Data!D{row}"
        ws[f"E{row}"] = f'=IF(A{row}="","",SUM(Input_Data!E{row}:H{row}))'
        ws[f"F{row}"] = f'=IF(E{row}="","",E{row}/4)'
        ws[f"G{row}"] = (
            f'=IF(E{row}="","",IF(COUNTA($E$2:$E$300)<=1,0,PERCENTRANK.INC($E$2:$E$300,E{row})))'
        )
        ws[f"H{row}"] = (
            f'=IF(G{row}="","",1-G{row})'
        )
        ws[f"I{row}"] = f"=Input_Data!I{row}"
        ws[f"J{row}"] = (
            f'=IF(I{row}="","",IF(COUNTA($I$2:$I$300)<=1,0,PERCENTRANK.INC($I$2:$I$300,I{row})))'
        )
        ws[f"K{row}"] = f"=Input_Data!J{row}"
        ws[f"L{row}"] = f"=Input_Data!K{row}"
        ws[f"M{row}"] = f"=Input_Data!L{row}"
        ws[f"N{row}"] = f'=IF(K{row}="","",K{row}-L{row})'
        ws[f"O{row}"] = f'=IF(A{row}="","",IF(ABS(M{row}-N{row})<=1,"OK","CHECK"))'
        ws[f"P{row}"] = f'=IF(M{row}="","",IF(F{row}=0,99,M{row}/F{row}))'
        ws[f"Q{row}"] = (
            f'=IF(P{row}="","",IF(COUNTA($P$2:$P$300)<=1,0,PERCENTRANK.INC($P$2:$P$300,P{row})))'
        )
        ws[f"R{row}"] = f'=IF(A{row}="","",0.45*H{row}+0.35*Q{row}+0.20*J{row})'
        ws[f"S{row}"] = (
            f'=IF(G{row}="","",IF(G{row}<=0.33,"Low Movement",IF(G{row}<=0.66,"Medium Movement","High Movement")))'
        )
        ws[f"T{row}"] = (
            f'=IF(R{row}="","",IF(R{row}>=0.67,"High Value Bundle",IF(R{row}>=0.34,"Medium Value Bundle","Low Value Bundle")))'
        )
        ws[f"U{row}"] = (
            f'=IF(A{row}="","",IF(AND(G{row}>=0.70,J{row}>=0.70,M{row}>0),"Yes","No"))'
        )
        ws[f"V{row}"] = (
            f'=IF(A{row}="","",IF(AND(S{row}="Low Movement",M{row}>0,P{row}>=2),"Yes","No"))'
        )
        ws[f"W{row}"] = (
            f'=IF(T{row}="","",MIN(0.12,IF(T{row}="High Value Bundle",0.05,IF(T{row}="Medium Value Bundle",0.07,0.10))+IF(P{row}>=8,0.02,0)))'
        )
        ws[f"X{row}"] = f'=IF(U{row}="Yes",C{row}&"|anchor","")'
        ws[f"Y{row}"] = f'=IF(U{row}="Yes",B{row},"")'

        for col in ("G", "H", "I", "J", "Q", "R", "W"):
            ws[f"{col}{row}"].number_format = "0.00%"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:Y300"
    set_widths(
        ws,
        {
            1: 14,
            2: 30,
            3: 20,
            4: 20,
            5: 16,
            6: 18,
            7: 18,
            8: 16,
            9: 16,
            10: 22,
            11: 12,
            12: 14,
            13: 14,
            14: 18,
            15: 15,
            16: 20,
            17: 23,
            18: 20,
            19: 18,
            20: 20,
            21: 15,
            22: 18,
            23: 20,
            24: 18,
            25: 25,
        },
    )

    ws.conditional_formatting.add(
        "T2:T300",
        FormulaRule(formula=['$T2="High Value Bundle"'], fill=PatternFill("solid", fgColor="FCE4D6")),
    )
    ws.conditional_formatting.add(
        "T2:T300",
        FormulaRule(formula=['$T2="Medium Value Bundle"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    ws.conditional_formatting.add(
        "T2:T300",
        FormulaRule(formula=['$T2="Low Value Bundle"'], fill=PatternFill("solid", fgColor="E2F0D9")),
    )
    ws.conditional_formatting.add(
        "O2:O300",
        FormulaRule(formula=['$O2="CHECK"'], fill=PatternFill("solid", fgColor="F8CBAD")),
    )


def build_recommendations_sheet(wb):
    ws = wb.create_sheet("Bundle_Recommendations")
    headers = [
        "product_id",
        "slow_mover_product",
        "category",
        "movement_band",
        "bundle_value_cluster",
        "priority_score",
        "stock_coverage_months",
        "recommended_anchor_product",
        "suggested_discount_%",
        "bundle_structure",
        "action_note",
    ]
    ws.append(headers)
    style_header(ws, ws[1])

    for row in range(2, MAX_ROWS):
        ws[f"A{row}"] = f"=Scoring_Model!A{row}"
        ws[f"B{row}"] = f"=Scoring_Model!B{row}"
        ws[f"C{row}"] = f"=Scoring_Model!C{row}"
        ws[f"D{row}"] = f"=Scoring_Model!S{row}"
        ws[f"E{row}"] = f"=Scoring_Model!T{row}"
        ws[f"F{row}"] = f"=Scoring_Model!R{row}"
        ws[f"G{row}"] = f"=Scoring_Model!P{row}"
        ws[f"H{row}"] = (
            f'=IF(Scoring_Model!V{row}<>"Yes","",IFERROR(XLOOKUP(C{row}&"|anchor",Scoring_Model!$X$2:$X$300,Scoring_Model!$Y$2:$Y$300,'
            f'INDEX(Scoring_Model!$Y$2:$Y$300,MATCH("Yes",Scoring_Model!$U$2:$U$300,0))),"No anchor found"))'
        )
        ws[f"I{row}"] = f"=Scoring_Model!W{row}"
        ws[f"J{row}"] = (
            f'=IF(Scoring_Model!V{row}<>"Yes","",IF(E{row}="High Value Bundle","1 anchor + 1 slow mover",IF(E{row}="Medium Value Bundle","1 anchor + 2 slow movers","1 anchor + 2 slow movers (aggressive clearout)")))'
        )
        ws[f"K{row}"] = (
            f'=IF(A{row}="","",IF(Scoring_Model!V{row}<>"Yes","Not a slow-mover bundle candidate","Bundle to move slow stock with controlled discount"))'
        )
        ws[f"F{row}"].number_format = "0.00%"
        ws[f"I{row}"].number_format = "0.00%"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:K300"
    set_widths(
        ws,
        {1: 14, 2: 32, 3: 20, 4: 18, 5: 20, 6: 14, 7: 21, 8: 30, 9: 18, 10: 34, 11: 52},
    )


def build_all_possible_bundles_placeholder(wb):
    ws = wb.create_sheet("All_Possible_Bundles")
    headers = [
        "bundle_id",
        "bundle_size",
        "anchor_product_id",
        "anchor_product_name",
        "item_2_product_id",
        "item_2_product_name",
        "item_3_product_id",
        "item_3_product_name",
        "category_mix",
        "bundle_priority_score",
        "suggested_discount_%",
        "reason",
    ]
    ws.append(headers)
    style_header(ws, ws[1])
    ws["A2"] = "Run script: python3 generate_all_possible_bundles.py bundle_planning_template.xlsx"
    ws["A3"] = "This fills all bundles with maximum 3 products (1 anchor + 1/2 slow movers)."
    ws.merge_cells("A2:L2")
    ws.merge_cells("A3:L3")
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="left")
    ws["A3"].alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L1"
    set_widths(
        ws,
        {
            1: 14,
            2: 12,
            3: 16,
            4: 30,
            5: 16,
            6: 30,
            7: 16,
            8: 30,
            9: 20,
            10: 20,
            11: 18,
            12: 48,
        },
    )


def build_logic_sheet(wb):
    ws = wb.create_sheet("Logic")
    ws["A1"] = "Bundle logic used in this workbook"
    ws["A1"].font = Font(bold=True, size=14)

    lines = [
        "1) Input columns follow your exact dataset: clean_product_name, category, product_id, purchased_item_count, purchased_qty_m1..m4, cont_from_total, stock, reserved_stock, available_stock.",
        "2) sold_qty_last_4m = SUM(purchased_qty_m1:purchased_qty_m4). Current month is intentionally excluded.",
        "3) Movement_Percentile ranks sold_qty_last_4m (0 = slowest mover, 1 = fastest mover).",
        "4) Slow_Mover_Score = 1 - Movement_Percentile (higher means slower mover, more urgent for bundling).",
        "5) Stock coverage = available_stock / avg_monthly_qty_4m. Higher coverage means higher stock pressure.",
        "6) Stock_Pressure_Percentile ranks stock_coverage_months to prioritize overstocked slow movers.",
        "7) Bundle_Priority_Score = 0.45*Slow_Mover_Score + 0.35*Stock_Pressure_Percentile + 0.20*Contribution_Percentile.",
        "8) Bundle clusters from Bundle_Priority_Score:",
        "   - High Value Bundle: score >= 0.67",
        "   - Medium Value Bundle: 0.34 <= score < 0.67",
        "   - Low Value Bundle: score < 0.34",
        "9) Anchor_Eligible = high movement + high contribution + available stock (>=70th percentile movement and contribution).",
        "10) Candidate_Eligible = Low Movement + available stock > 0 + stock coverage >=2 months.",
        "11) Recommended bundle design: every candidate includes one anchor product that attracts customers.",
        "12) Suggested discount is intentionally slight to control burn:",
        "    - High Value: 5%",
        "    - Medium Value: 7%",
        "    - Low Value: 10%",
        "    - Add +2% only when stock coverage >=8 months (capped at 12%).",
        "13) stock_data_check flags rows where available_stock does not match stock - reserved_stock.",
        "14) Use generate_all_possible_bundles.py to produce all bundle combinations up to 3 products.",
    ]

    row = 3
    for line in lines:
        ws[f"A{row}"] = line
        row += 1

    ws.column_dimensions["A"].width = 150


def main():
    wb = Workbook()
    build_input_sheet(wb)
    build_scoring_sheet(wb)
    build_recommendations_sheet(wb)
    build_all_possible_bundles_placeholder(wb)
    build_logic_sheet(wb)
    wb.save(OUTPUT_FILE)
    print(f"Workbook created: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
