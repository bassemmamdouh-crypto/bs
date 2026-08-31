import math
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from area_item_target_planner import (
    PlannerError,
    build_mix_and_targets,
    last_n_months,
    load_purchasing_table,
    normalize_purchasing_frame,
    parse_month_label,
    plan_from_excel,
    write_output_workbook,
)
from create_sample_purchasing_workbook import create_sample


def test_parse_month_label_variants():
    assert parse_month_label("2026-05") == pd.Period("2026-05", "M")
    assert parse_month_label("May 2026") == pd.Period("2026-05", "M")
    assert parse_month_label(pd.Timestamp("2026-05-17")) == pd.Period("2026-05", "M")
    assert parse_month_label("item") is None


def test_last_three_months_are_calendar_window():
    assert last_n_months(pd.Period("2026-07", "M"), 3) == [
        pd.Period("2026-05", "M"),
        pd.Period("2026-06", "M"),
        pd.Period("2026-07", "M"),
    ]


def _long_frame():
    rows = [
        ("Baghdad", "P330", "Pepsi 330ml", "2026-04", 999999, 1),  # outside window
        ("Baghdad", "P330", "Pepsi 330ml", "2026-05", 100, 10),
        ("Baghdad", "P330", "Pepsi 330ml", "2026-06", 100, 10),
        ("Baghdad", "P330", "Pepsi 330ml", "2026-07", 100, 10),
        ("Baghdad", "L25", "Lays 25g", "2026-05", 100, 20),
        ("Baghdad", "L25", "Lays 25g", "2026-06", 300, 60),
        ("Baghdad", "L25", "Lays 25g", "2026-07", 200, 40),
        ("Basra", "P330", "Pepsi 330ml", "2026-05", 50, 5),
        ("Basra", "P330", "Pepsi 330ml", "2026-06", 50, 5),
        ("Basra", "P330", "Pepsi 330ml", "2026-07", 50, 5),
    ]
    return pd.DataFrame(rows, columns=["area", "sku", "item", "month", "sales", "qty"])


def test_weighted_contribution_and_targets_sum_to_area_total():
    purchasing = normalize_purchasing_frame(_long_frame())
    result = build_mix_and_targets(purchasing, lookback_months=3)
    items = result["item_targets"]
    baghdad = items[items["area"] == "Baghdad"].set_index("sku")

    # Last 3 months Baghdad: Pepsi 300, Lays 600, total 900 → 1/3 and 2/3.
    assert baghdad.loc["P330", "weighted_contribution"] == pytest.approx(300 / 900)
    assert baghdad.loc["L25", "weighted_contribution"] == pytest.approx(600 / 900)
    # Simple average of monthly shares: Pepsi 50%, 25%, 33.3% → ~36.11%
    assert baghdad.loc["P330", "simple_avg_contribution"] == pytest.approx(
        (0.5 + 0.25 + 1 / 3) / 3
    )
    assert result["target_month"] == "2026-08"

    area_default = result["area_summary"].set_index("area").loc["Baghdad", "next_month_target"]
    assert area_default == pytest.approx(900 / 3)
    assert items[items["area"] == "Baghdad"]["next_month_sales_target"].sum() == pytest.approx(
        area_default
    )
    # April sales must not leak into mix.
    assert baghdad.loc["P330", "sales_last_n"] == pytest.approx(300)


def test_area_target_override_scales_items():
    purchasing = normalize_purchasing_frame(_long_frame())
    overrides = pd.DataFrame({"area": ["Baghdad"], "next_month_target": [1200.0]})
    result = build_mix_and_targets(purchasing, area_targets=overrides)
    baghdad = result["item_targets"]
    baghdad = baghdad[baghdad["area"] == "Baghdad"]
    assert baghdad["next_month_sales_target"].sum() == pytest.approx(1200)
    pepsi = baghdad.set_index("sku").loc["P330", "next_month_sales_target"]
    assert pepsi == pytest.approx(1200 * (300 / 900))
    source = result["area_summary"].set_index("area").loc["Baghdad", "target_source"]
    assert source == "override"


def test_qty_target_uses_implied_unit_price():
    purchasing = normalize_purchasing_frame(_long_frame())
    result = build_mix_and_targets(purchasing)
    pepsi = result["item_targets"].set_index(["area", "sku"]).loc[("Baghdad", "P330")]
    assert pepsi["unit_price"] == pytest.approx(10.0)
    assert pepsi["next_month_qty_target"] == pytest.approx(pepsi["next_month_sales_target"] / 10.0)


def test_wide_format_matches_long_format(tmp_path: Path):
    sample = create_sample(tmp_path / "sample.xlsx")
    wide = load_purchasing_table(sample, sheet="Purchasing_Wide")
    long = load_purchasing_table(sample, sheet="Purchasing_Long")
    wide_sales = wide.groupby(["area", "sku", "month"])["sales"].sum()
    long_sales = long.groupby(["area", "sku", "month"])["sales"].sum()
    pd.testing.assert_series_equal(wide_sales.sort_index(), long_sales.sort_index())


def test_duplicate_rows_are_summed():
    raw = pd.DataFrame(
        [
            ("Baghdad", "A", "Item A", "2026-07", 10),
            ("Baghdad", "A", "Item A", "2026-07", 15),
            ("Baghdad", "B", "Item B", "2026-07", 75),
        ],
        columns=["area", "sku", "item", "month", "sales"],
    )
    frame = normalize_purchasing_frame(raw)
    a = frame[(frame["sku"] == "A") & (frame["month"] == pd.Period("2026-07", "M"))]
    assert a["sales"].iloc[0] == 25


def test_end_to_end_writes_expected_sheets(tmp_path: Path):
    sample = create_sample(tmp_path / "sample.xlsx")
    output = tmp_path / "targets.xlsx"
    result = plan_from_excel(sample, output, sheet="Purchasing_Long")
    assert output.exists()
    wb = load_workbook(output)
    for name in ("Instructions", "Area_Targets", "Item_Targets", "Monthly_Mix", "Baghdad", "Basra", "Erbil"):
        assert name in wb.sheetnames

    items = result["item_targets"]
    for area, group in items.groupby("area"):
        area_total = result["area_summary"].set_index("area").loc[area, "next_month_target"]
        assert group["weighted_contribution"].sum() == pytest.approx(1.0)
        assert group["next_month_sales_target"].sum() == pytest.approx(area_total)

    # Baghdad mix on sample data: Pepsi 330 = 375000 / 795000
    baghdad = items[items["area"] == "Baghdad"].set_index("sku")
    assert baghdad.loc["P330", "weighted_contribution"] == pytest.approx(375000 / 795000)
    assert math.isclose(baghdad["weighted_contribution"].sum(), 1.0)


def test_missing_area_column_raises():
    raw = pd.DataFrame({"item": ["A"], "month": ["2026-07"], "sales": [1]})
    with pytest.raises(PlannerError, match="area column"):
        normalize_purchasing_frame(raw)


def test_arabic_month_headers_and_two_month_average():
    raw = pd.DataFrame(
        {
            "المنطقة": ["الحرية", "الحرية", "الحرية", "شعلة"],
            "اسم المشرف / الفريق": ["رسول", "رسول", "رسول", "ميسر"],
            "مركز الكلفة": ["احمد سعدي", "احمد سعدي", "موزع 18", "موزع 3"],
            "العبوة (Container)": ["1.25 ltr", "1.25 ltr", "Can185 ml", "1.25 ltr"],
            "الصنف (ROUT)": ["Pepsi", "7-Up", "Pepsi", "Pepsi"],
            "شهر 6": [300, 100, 100, 50],
            "شهر 7": [500, 100, 0, 50],
            "الإجمالي": [800, 200, 100, 100],
        }
    )
    frame = normalize_purchasing_frame(raw, default_year=2026)
    assert set(frame["month"].astype(str)) == {"2026-06", "2026-07"}
    result = build_mix_and_targets(frame, lookback_months=3)
    assert result["target_month"] == "2026-08"
    assert result["months_in_file"] == 2
    hurriya = result["item_targets"]
    hurriya = hurriya[hurriya["area"] == "الحرية"].set_index("sku")
    # Area last-2-month total = 300+100+100 + 500+100+0 = 1100, default target = 550
    assert hurriya.loc["1.25 ltr | Pepsi", "weighted_contribution"] == pytest.approx(800 / 1100)
    assert hurriya["next_month_sales_target"].sum() == pytest.approx(550)
    cc = result["cost_center_targets"]
    pepsi_125 = cc[(cc["sku"] == "1.25 ltr | Pepsi") & (cc["area"] == "الحرية")]
    assert pepsi_125["next_month_sales_target"].sum() == pytest.approx(
        hurriya.loc["1.25 ltr | Pepsi", "next_month_sales_target"]
    )


def test_parse_arabic_month_header():
    assert parse_month_label("شهر 6", default_year=2026) == pd.Period("2026-06", "M")
    assert parse_month_label("الإجمالي", default_year=2026) is None


def test_cli_main(tmp_path: Path):
    sample = create_sample(tmp_path / "sample.xlsx")
    output = tmp_path / "out.xlsx"
    from area_item_target_planner import main

    assert main([str(sample), "-o", str(output), "--sheet", "Purchasing_Wide"]) == 0
    assert output.exists()
    wb = load_workbook(output)
    assert wb["Item_Targets"]["A1"].value == "area"
