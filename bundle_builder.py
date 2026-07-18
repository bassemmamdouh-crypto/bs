import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MAX_PRODUCTS = 300
INPUT_SHEET = "Input_Data"
SCORING_SHEET = "Scoring_Model"
BUNDLE_SHEET = "All_Possible_Bundles"
TOP_BUNDLE_SHEET = "Top_30_Bundles"

# KEEP ONLY TOP N STRONGEST PRODUCTS AS ANCHORS
MAX_ANCHORS = 7

# KEEP ONLY THE TOP N BUNDLES (BY BUNDLE PRIORITY SCORE)
TOP_N_BUNDLES = 30

# Prefer same-category anchors when choosing the single best anchor
SAME_CATEGORY_BONUS = 0.10

# Soft penalty so movers spread across anchors instead of all
# locking onto the single strongest product
ANCHOR_LOAD_PENALTY = 0.03

# Bundle composition:
#   exactly 1 anchor + 2 other products (always 3-product bundles)
#   at most 1 medium mover OR 1 high mover (not both)
#   high movers may only share a bundle with slow movers
MAX_MEDIUM_MOVERS_PER_BUNDLE = 1
MAX_HIGH_MOVERS_PER_BUNDLE = 1
MAX_SLOW_MOVERS_PER_BUNDLE = 2
NON_ANCHOR_ITEMS_PER_BUNDLE = 2

# Suggested discount ceiling
MAX_DISCOUNT = 0.05


def to_number(value, default=0.0):
    if value is None or value == "":
        return default

    try:
        return float(value)

    except (TypeError, ValueError):
        return default


def normalize_percent(value):

    number = to_number(value, 0.0)

    return number / 100.0 if number > 1 else number


def percent_rank_inc(values, target):

    sorted_vals = sorted(values)

    n = len(sorted_vals)

    if n <= 1:
        return 0.0

    if target <= sorted_vals[0]:
        return 0.0

    if target >= sorted_vals[-1]:
        return 1.0

    for idx, val in enumerate(sorted_vals):

        if val == target:

            first = idx
            last = idx

            while first - 1 >= 0 and sorted_vals[first - 1] == target:
                first -= 1

            while last + 1 < n and sorted_vals[last + 1] == target:
                last += 1

            return ((first + last) / 2.0) / (n - 1)

    for idx in range(1, n):

        lower = sorted_vals[idx - 1]
        upper = sorted_vals[idx]

        if lower < target < upper:

            frac = (target - lower) / (upper - lower)

            return ((idx - 1) + frac) / (n - 1)

    return 0.0


def read_products(ws):

    products = []

    for row in range(2, MAX_PRODUCTS + 2):

        name = ws[f"A{row}"].value
        category = ws[f"B{row}"].value
        product_id = ws[f"C{row}"].value

        if not product_id:
            continue

        qty_m1 = to_number(ws[f"E{row}"].value)
        qty_m2 = to_number(ws[f"F{row}"].value)
        qty_m3 = to_number(ws[f"G{row}"].value)
        qty_m4 = to_number(ws[f"H{row}"].value)

        sold_qty_4m = qty_m1 + qty_m2 + qty_m3 + qty_m4

        avg_qty_4m = (
            sold_qty_4m / 4
            if sold_qty_4m > 0
            else 0.0
        )

        stock = to_number(ws[f"J{row}"].value)

        reserved_stock = to_number(ws[f"K{row}"].value)

        available_stock = to_number(
            ws[f"L{row}"].value,
            stock - reserved_stock
        )

        if available_stock == 0 and stock > 0:
            available_stock = stock - reserved_stock

        products.append(
            {
                "product_id": str(product_id),
                "name": str(name) if name else "",
                "category": str(category) if category else "",

                "purchased_item_count": to_number(
                    ws[f"D{row}"].value
                ),

                "sold_qty_last_4m": sold_qty_4m,

                "avg_qty_4m": avg_qty_4m,

                "contribution": normalize_percent(
                    ws[f"I{row}"].value
                ),

                "stock": stock,

                "reserved_stock": reserved_stock,

                "available_stock": max(
                    available_stock,
                    0.0
                ),
            }
        )

    return products


def score_products(products):

    sold_values = [
        p["sold_qty_last_4m"]
        for p in products
    ]

    contribution_values = [
        p["contribution"]
        for p in products
    ]

    coverage_values = []

    for product in products:

        avg = product["avg_qty_4m"]

        coverage = (
            99.0
            if avg == 0
            else product["available_stock"] / avg
        )

        coverage_values.append(coverage)

    for idx, product in enumerate(products):

        movement_percentile = percent_rank_inc(
            sold_values,
            product["sold_qty_last_4m"]
        )

        slow_mover_score = 1.0 - movement_percentile

        contribution_percentile = percent_rank_inc(
            contribution_values,
            product["contribution"]
        )

        stock_coverage = coverage_values[idx]

        stock_pressure_percentile = percent_rank_inc(
            coverage_values,
            stock_coverage
        )

        priority_score = (
            0.45 * slow_mover_score
            + 0.35 * stock_pressure_percentile
            + 0.20 * contribution_percentile
        )

        # MOVEMENT BAND
        if movement_percentile <= 0.33:
            movement_band = "Low Movement"

        elif movement_percentile <= 0.66:
            movement_band = "Medium Movement"

        else:
            movement_band = "High Movement"

        # CLUSTER
        if priority_score >= 0.67:

            cluster = "High Value Bundle"

            base_discount = 0.03

        elif priority_score >= 0.34:

            cluster = "Medium Value Bundle"

            base_discount = 0.04

        else:

            cluster = "Low Value Bundle"

            base_discount = 0.05

        discount = min(
            MAX_DISCOUNT,
            base_discount + (
                0.02 if stock_coverage >= 8 else 0.0
            )
        )

        # STRONGEST PRODUCTS ONLY
        anchor_eligible = (
            movement_percentile >= 0.85
            and contribution_percentile >= 0.70
            and product["available_stock"] > 0
        )

        anchor_strength = (
            0.6 * movement_percentile
            + 0.4 * contribution_percentile
        )

        product.update(
            {
                "movement_percentile": movement_percentile,

                "slow_mover_score": slow_mover_score,

                "contribution_percentile": contribution_percentile,

                "stock_coverage": stock_coverage,

                "stock_pressure_percentile": stock_pressure_percentile,

                "priority_score": priority_score,

                "movement_band": movement_band,

                "cluster": cluster,

                "discount": discount,

                "anchor_eligible": anchor_eligible,

                "anchor_strength": anchor_strength,
            }
        )

    # KEEP ONLY TOP ANCHORS
    sorted_anchors = sorted(
        [p for p in products if p["anchor_eligible"]],
        key=lambda x: x["anchor_strength"],
        reverse=True
    )

    top_anchor_ids = {
        p["product_id"]
        for p in sorted_anchors[:MAX_ANCHORS]
    }

    # FINALIZE ROLES: anchor / high / medium / slow mover
    for product in products:

        product["anchor_eligible"] = (
            product["product_id"] in top_anchor_ids
        )

        if product["anchor_eligible"]:
            mover_type = "anchor"

        elif product["movement_band"] == "Low Movement":
            mover_type = "slow_mover"

        elif product["movement_band"] == "Medium Movement":
            mover_type = "medium_mover"

        else:
            # High movers that are not selected as anchors
            mover_type = "high_mover"

        product["mover_type"] = mover_type

        # High, medium, and slow movers can fill bundles with an anchor
        product["candidate_eligible"] = (
            mover_type in {
                "high_mover",
                "medium_mover",
                "slow_mover",
            }
            and product["available_stock"] > 0
        )


def update_scoring_sheet(ws, products):

    product_map = {
        p["product_id"]: p
        for p in products
    }

    for row in range(2, MAX_PRODUCTS + 2):

        product_id = ws[f"A{row}"].value

        if not product_id:
            continue

        product = product_map.get(str(product_id))

        if not product:
            continue

        ws[f"E{row}"] = product["sold_qty_last_4m"]
        ws[f"F{row}"] = product["avg_qty_4m"]
        ws[f"G{row}"] = product["movement_percentile"]
        ws[f"H{row}"] = product["slow_mover_score"]
        ws[f"I{row}"] = product["contribution"]
        ws[f"J{row}"] = product["contribution_percentile"]
        ws[f"K{row}"] = product["stock"]
        ws[f"L{row}"] = product["reserved_stock"]
        ws[f"M{row}"] = product["available_stock"]

        ws[f"N{row}"] = (
            product["stock"]
            - product["reserved_stock"]
        )

        ws[f"O{row}"] = (
            "OK"
            if abs(
                product["available_stock"]
                - (
                    product["stock"]
                    - product["reserved_stock"]
                )
            ) <= 1
            else "CHECK"
        )

        ws[f"P{row}"] = product["stock_coverage"]
        ws[f"Q{row}"] = product["stock_pressure_percentile"]
        ws[f"R{row}"] = product["priority_score"]
        ws[f"S{row}"] = product["movement_band"]
        ws[f"T{row}"] = product["cluster"]

        ws[f"U{row}"] = (
            "Yes"
            if product["anchor_eligible"]
            else "No"
        )

        ws[f"V{row}"] = (
            "Yes"
            if product["candidate_eligible"]
            else "No"
        )

        ws[f"W{row}"] = product["discount"]

        ws[f"X{row}"] = (
            f'{product["category"]}|anchor'
            if product["anchor_eligible"]
            else ""
        )

        ws[f"Y{row}"] = (
            product["name"]
            if product["anchor_eligible"]
            else ""
        )

        for col in (
            "G",
            "H",
            "I",
            "J",
            "Q",
            "R",
            "W"
        ):
            ws[f"{col}{row}"].number_format = "0.00%"


def style_header(ws):

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    for cell in ws[1]:

        cell.font = Font(
            color="FFFFFF",
            bold=True
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


def set_widths(ws, widths):

    for col_idx, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col_idx)
        ].width = width


def pair_score(movers, anchor):
    """Score a set of non-anchor movers against one anchor."""

    if not movers:
        return 0.0

    ordered = sorted(
        movers,
        key=lambda p: p["priority_score"],
        reverse=True
    )

    if len(ordered) == 1:
        score = (
            0.75 * ordered[0]["priority_score"]
            + 0.25 * anchor["anchor_strength"]
        )
    else:
        score = (
            0.45 * ordered[0]["priority_score"]
            + 0.35 * ordered[1]["priority_score"]
            + 0.20 * anchor["anchor_strength"]
        )

    same_category = all(
        mover["category"] == anchor["category"]
        for mover in ordered
        if mover["category"]
    )

    if same_category:
        score += SAME_CATEGORY_BONUS

    return score


def choose_best_anchor(mover, anchors, anchor_loads):
    """Pick one anchor for a mover (category + strength + load balance)."""

    best_anchor = None
    best_adjusted = None

    for anchor in anchors:

        if anchor["product_id"] == mover["product_id"]:
            continue

        score = pair_score([mover], anchor)

        adjusted = (
            score
            - ANCHOR_LOAD_PENALTY
            * anchor_loads[anchor["product_id"]]
        )

        if best_adjusted is None or adjusted > best_adjusted:
            best_adjusted = adjusted
            best_anchor = anchor

    return best_anchor


def category_mix_label(anchor, movers):
    categories = {
        mover["category"]
        for mover in movers
        if mover["category"]
    }

    if not categories:
        return "Cross Category"

    if (
        len(categories) == 1
        and anchor["category"] in categories
    ):
        return "Same Category"

    if len(movers) == 1:
        return "Cross Category"

    return "Mixed Category"


def bundle_reason(movers):
    high_count = sum(
        1 for m in movers if m["mover_type"] == "high_mover"
    )
    medium_count = sum(
        1 for m in movers if m["mover_type"] == "medium_mover"
    )
    slow_count = sum(
        1 for m in movers if m["mover_type"] == "slow_mover"
    )

    if high_count and slow_count:
        return "One anchor + one high mover + one slow mover"

    if medium_count and slow_count:
        return (
            "One anchor + one medium mover + one slow mover"
        )

    return "One anchor + two slow movers"


def make_bundle(anchor, movers):
    """Build one 3-product bundle: 1 anchor + 2 movers."""

    if len(movers) != NON_ANCHOR_ITEMS_PER_BUNDLE:
        raise ValueError(
            "Every bundle must have exactly 3 products"
        )

    high_count = sum(
        1 for m in movers if m["mover_type"] == "high_mover"
    )
    medium_count = sum(
        1 for m in movers if m["mover_type"] == "medium_mover"
    )
    slow_count = sum(
        1 for m in movers if m["mover_type"] == "slow_mover"
    )

    if high_count > MAX_HIGH_MOVERS_PER_BUNDLE:
        raise ValueError("Bundle exceeds max high movers")

    if medium_count > MAX_MEDIUM_MOVERS_PER_BUNDLE:
        raise ValueError("Bundle exceeds max medium movers")

    if slow_count > MAX_SLOW_MOVERS_PER_BUNDLE:
        raise ValueError("Bundle exceeds max slow movers")

    if high_count and medium_count:
        raise ValueError(
            "High movers cannot share a bundle with medium movers"
        )

    if high_count and slow_count == 0:
        raise ValueError(
            "High movers must be bundled with slow movers"
        )

    if medium_count and slow_count == 0:
        raise ValueError(
            "Medium movers must be bundled with slow movers"
        )

    # Item order: high/medium first, then slow movers
    type_rank = {
        "high_mover": 0,
        "medium_mover": 0,
        "slow_mover": 1,
    }

    ordered = sorted(
        movers,
        key=lambda m: (
            type_rank.get(m["mover_type"], 9),
            -m["priority_score"],
        ),
    )

    item_2 = ordered[0]
    item_3 = ordered[1]

    score = pair_score(ordered, anchor)
    discount = min(
        MAX_DISCOUNT,
        max(m["discount"] for m in ordered),
    )

    return {
        "score": score,
        "anchor_id": anchor["product_id"],
        "candidate_ids": {m["product_id"] for m in ordered},
        "row": [
            3,

            anchor["product_id"],
            anchor["name"],

            item_2["product_id"],
            item_2["name"],

            item_3["product_id"],
            item_3["name"],

            category_mix_label(anchor, ordered),

            score,

            discount,

            bundle_reason(ordered),
        ],
    }


def pack_anchor_bundles(
    anchor,
    high_movers,
    medium_movers,
    slow_movers,
):
    """Pack one anchor's assigned movers into 3-product bundles only.

    Rules:
    - exactly 1 anchor + 2 other products
    - high movers only with slow movers (never with medium)
    - medium movers only with slow movers
    - leftover slows pair as two slow movers

    Shapes produced:
    1. anchor + 1 high + 1 slow
    2. anchor + 1 medium + 1 slow
    3. anchor + 2 slow
    """

    highs = sorted(
        high_movers,
        key=lambda p: p["priority_score"],
        reverse=True,
    )
    mediums = sorted(
        medium_movers,
        key=lambda p: p["priority_score"],
        reverse=True,
    )
    slows = sorted(
        slow_movers,
        key=lambda p: p["priority_score"],
        reverse=True,
    )

    bundles = []

    # High movers must ride with slow movers only
    while highs and slows:
        bundles.append(
            make_bundle(
                anchor,
                [highs.pop(0), slows.pop(0)],
            )
        )

    # Medium movers must include a slow mover
    while mediums and slows:
        bundles.append(
            make_bundle(
                anchor,
                [mediums.pop(0), slows.pop(0)],
            )
        )

    # Leftover slows: anchor + two slow movers
    while len(slows) >= 2:
        bundles.append(
            make_bundle(
                anchor,
                [slows.pop(0), slows.pop(0)],
            )
        )

    # Unpaired high/medium/single slow are skipped
    # (cannot form a valid 3-product bundle)
    return bundles


def build_candidate_bundles(products):
    """Build 3-product bundles from anchors + high/medium/slow movers.

    Each non-anchor mover is assigned to exactly one anchor, then packed
    into bundles with:
    - exactly 3 products (1 anchor + 2 movers)
    - high/medium movers only alongside slow movers
    """

    anchors = [
        p for p in products
        if p["anchor_eligible"]
    ]

    high_movers = [
        p for p in products
        if p["candidate_eligible"]
        and p["mover_type"] == "high_mover"
    ]

    medium_movers = [
        p for p in products
        if p["candidate_eligible"]
        and p["mover_type"] == "medium_mover"
    ]

    slow_movers = [
        p for p in products
        if p["candidate_eligible"]
        and p["mover_type"] == "slow_mover"
    ]

    anchors.sort(
        key=lambda x: x["anchor_strength"],
        reverse=True
    )

    if not anchors:
        return []

    assignments = {
        anchor["product_id"]: {
            "anchor": anchor,
            "high": [],
            "medium": [],
            "slow": [],
        }
        for anchor in anchors
    }

    anchor_loads = {
        anchor["product_id"]: 0
        for anchor in anchors
    }

    def slow_partner_need(anchor_id):
        bucket = assignments[anchor_id]
        return max(
            0,
            len(bucket["high"])
            + len(bucket["medium"])
            - len(bucket["slow"]),
        )

    def choose_anchor_for_slow(mover):
        """Prefer anchors that still need a slow partner for high/medium."""

        needing = [
            a for a in anchors
            if slow_partner_need(a["product_id"]) > 0
            and a["product_id"] != mover["product_id"]
        ]

        pool = needing or [
            a for a in anchors
            if a["product_id"] != mover["product_id"]
        ]

        best_anchor = None
        best_adjusted = None

        for anchor in pool:
            score = pair_score([mover], anchor)
            adjusted = (
                score
                - ANCHOR_LOAD_PENALTY
                * anchor_loads[anchor["product_id"]]
            )

            if best_adjusted is None or adjusted > best_adjusted:
                best_adjusted = adjusted
                best_anchor = anchor

        return best_anchor

    # Assign high/medium first so slows can fill their partner slots
    lead_movers = sorted(
        high_movers + medium_movers,
        key=lambda p: p["priority_score"],
        reverse=True,
    )

    for mover in lead_movers:

        best_anchor = choose_best_anchor(
            mover,
            anchors,
            anchor_loads,
        )

        if best_anchor is None:
            continue

        bucket = assignments[best_anchor["product_id"]]

        if mover["mover_type"] == "high_mover":
            bucket["high"].append(mover)
        else:
            bucket["medium"].append(mover)

        anchor_loads[best_anchor["product_id"]] += 1

    for mover in sorted(
        slow_movers,
        key=lambda p: p["priority_score"],
        reverse=True,
    ):

        best_anchor = choose_anchor_for_slow(mover)

        if best_anchor is None:
            continue

        assignments[best_anchor["product_id"]]["slow"].append(
            mover
        )
        anchor_loads[best_anchor["product_id"]] += 1

    bundles = []

    for anchor in anchors:

        bucket = assignments[anchor["product_id"]]

        bundles.extend(
            pack_anchor_bundles(
                anchor,
                bucket["high"],
                bucket["medium"],
                bucket["slow"],
            )
        )

    return bundles


def write_bundle_sheet(ws, bundles):
    """Write the given (already ranked) bundles to a worksheet.

    ``bundles`` is a list of dicts as produced by ``build_candidate_bundles``.
    Rows are re-numbered sequentially (``B-000001``, ``B-000002``, ...).
    """

    headers = [
        "bundle_id",
        "bundle_size",
        "anchor_product_id",
        "anchor_product_name",
        "item_2_product_id",
        "item_2_product_name",
        "item_3_product_id",
        "item_3_product_name",
        "category_mix",
        "bundle_priority_score",
        "suggested_discount_%",
        "reason",
    ]

    merged_ranges = list(
        ws.merged_cells.ranges
    )

    for cell_range in merged_ranges:
        ws.unmerge_cells(str(cell_range))

    ws.delete_rows(1, ws.max_row)

    ws.append(headers)

    style_header(ws)

    for bundle_id, bundle in enumerate(bundles, start=1):

        ws.append(
            [f"B-{bundle_id:06d}"] + bundle["row"]
        )

    for row in range(2, ws.max_row + 1):

        ws[f"J{row}"].number_format = "0.00%"

        ws[f"K{row}"].number_format = "0.00%"

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    set_widths(
        ws,
        {
            1: 14,
            2: 12,
            3: 16,
            4: 30,
            5: 16,
            6: 30,
            7: 16,
            8: 30,
            9: 20,
            10: 20,
            11: 18,
            12: 56,
        },
    )

    return len(bundles)


def ensure_sheet(workbook, name):

    if name in workbook.sheetnames:
        return workbook[name]

    return workbook.create_sheet(name)


def main():

    parser = argparse.ArgumentParser(
        description="Generate the top bundles"
    )

    parser.add_argument(
        "workbook_path",
        nargs="?",
        default="bundle_planning_template.xlsx",
        help="Path to workbook"
    )

    # FIX FOR JUPYTER
    args, unknown = parser.parse_known_args()

    workbook_path = Path(args.workbook_path)

    print(f"Loading workbook: {workbook_path}")

    wb = load_workbook(workbook_path)

    input_ws = wb[INPUT_SHEET]

    products = read_products(input_ws)

    if not products:
        raise ValueError(
            "No products found in Input_Data"
        )

    print(f"Products loaded: {len(products)}")

    score_products(products)

    scoring_ws = ensure_sheet(
        wb,
        SCORING_SHEET
    )

    update_scoring_sheet(
        scoring_ws,
        products
    )

    # 1 ANCHOR + MAX 1 MEDIUM MOVER + SLOW MOVERS
    bundles = build_candidate_bundles(products)

    bundles.sort(
        key=lambda b: b["score"],
        reverse=True
    )

    top_bundles = bundles[:TOP_N_BUNDLES]

    # TAB 1: ALL PACKED BUNDLES
    all_bundles_ws = ensure_sheet(
        wb,
        BUNDLE_SHEET
    )

    total_bundles = write_bundle_sheet(
        all_bundles_ws,
        bundles
    )

    # TAB 2: TOP N BUNDLES
    top_bundles_ws = ensure_sheet(
        wb,
        TOP_BUNDLE_SHEET
    )

    kept_bundles = write_bundle_sheet(
        top_bundles_ws,
        top_bundles
    )

    wb.save(workbook_path)

    print("\nWorkbook updated successfully.")

    print(
        f"Saved to: "
        f"{workbook_path.resolve()}"
    )

    print(f"Products: {len(products)}")

    print(
        f"Anchors: "
        f"{sum(1 for p in products if p['anchor_eligible'])}"
    )

    print(
        f"High movers: "
        f"{sum(1 for p in products if p.get('mover_type') == 'high_mover' and p['candidate_eligible'])}"
    )

    print(
        f"Medium movers: "
        f"{sum(1 for p in products if p.get('mover_type') == 'medium_mover' and p['candidate_eligible'])}"
    )

    print(
        f"Slow movers: "
        f"{sum(1 for p in products if p.get('mover_type') == 'slow_mover' and p['candidate_eligible'])}"
    )

    print(
        f"'{BUNDLE_SHEET}' tab: {total_bundles} bundles "
        f"(3-product bundles; discount capped at {MAX_DISCOUNT:.0%})"
    )

    print(
        f"'{TOP_BUNDLE_SHEET}' tab: {kept_bundles} bundles "
        f"(top {TOP_N_BUNDLES} by priority score)"
    )


if __name__ == "__main__":
    main()
