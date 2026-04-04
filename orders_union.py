#!/usr/bin/env python3
"""
Union orders data from multiple sheets into one target sheet.

Example:
  python orders_union.py --input orders.xlsx --target-sheet "All Orders"
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def parse_csv_names(value: str | None) -> set[str]:
    if not value:
        return set()
    return {item.strip() for item in value.split(",") if item.strip()}


def is_blank_row(values: Iterable[object]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in values)


def normalize_header(value: object, index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text if text else f"Column {index + 1}"


def should_include_sheet(
    sheet_name: str,
    *,
    target_sheet_name: str,
    include_sheets: set[str],
    exclude_sheets: set[str],
    source_name_prefix: str,
) -> bool:
    if sheet_name == target_sheet_name:
        return False
    if sheet_name in exclude_sheets:
        return False
    if include_sheets and sheet_name not in include_sheets:
        return False
    if source_name_prefix and not sheet_name.startswith(source_name_prefix):
        return False
    return True


def union_orders_data(
    input_path: Path,
    output_path: Path | None = None,
    *,
    target_sheet_name: str = "All Orders",
    header_row: int = 1,
    include_sheets: set[str] | None = None,
    exclude_sheets: set[str] | None = None,
    source_name_prefix: str = "",
    add_source_sheet_column: bool = True,
    source_sheet_column_name: str = "Source Sheet",
    clear_target_before_write: bool = True,
) -> tuple[int, int]:
    if header_row < 1:
        raise ValueError("header_row must be >= 1")

    include_sheets = include_sheets or set()
    exclude_sheets = exclude_sheets or set()

    workbook = load_workbook(input_path)

    source_sheets = [
        sheet
        for sheet in workbook.worksheets
        if should_include_sheet(
            sheet.title,
            target_sheet_name=target_sheet_name,
            include_sheets=include_sheets,
            exclude_sheets=exclude_sheets,
            source_name_prefix=source_name_prefix,
        )
    ]

    if not source_sheets:
        raise ValueError("No source sheets found with the current filters.")

    master_headers: list[str] = []
    master_headers_set: set[str] = set()
    row_objects: list[dict[str, object]] = []

    for sheet in source_sheets:
        max_col = sheet.max_column
        max_row = sheet.max_row
        if max_col < 1 or max_row < header_row:
            continue

        raw_headers = [sheet.cell(row=header_row, column=col).value for col in range(1, max_col + 1)]
        normalized_headers = [normalize_header(value, idx) for idx, value in enumerate(raw_headers)]

        for header in normalized_headers:
            if header not in master_headers_set:
                master_headers_set.add(header)
                master_headers.append(header)

        for row_idx in range(header_row + 1, max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
            if is_blank_row(row_values):
                continue

            row_object: dict[str, object] = {}
            for col_idx, header in enumerate(normalized_headers):
                row_object[header] = row_values[col_idx] if col_idx < len(row_values) else None

            if add_source_sheet_column:
                row_object[source_sheet_column_name] = sheet.title

            row_objects.append(row_object)

    if add_source_sheet_column and source_sheet_column_name not in master_headers_set:
        master_headers.append(source_sheet_column_name)

    if target_sheet_name in workbook.sheetnames:
        target_sheet = workbook[target_sheet_name]
    else:
        target_sheet = workbook.create_sheet(title=target_sheet_name)

    if clear_target_before_write and target_sheet.max_row > 0:
        target_sheet.delete_rows(1, target_sheet.max_row)

    if not master_headers:
        target_sheet.cell(row=1, column=1, value="No data found.")
    else:
        target_sheet.append(master_headers)
        for row_object in row_objects:
            target_sheet.append([row_object.get(header, None) for header in master_headers])

    save_path = output_path or input_path
    workbook.save(save_path)

    return len(row_objects), len(master_headers)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Scan orders from multiple sheets and union them into one sheet."
    )
    parser.add_argument("--input", required=True, help="Path to input XLSX file.")
    parser.add_argument(
        "--output",
        help="Optional output XLSX path. If omitted, input file is updated in place.",
    )
    parser.add_argument("--target-sheet", default="All Orders", help="Output sheet name.")
    parser.add_argument("--header-row", type=int, default=1, help="Header row number (1-based).")
    parser.add_argument(
        "--include",
        help="Comma-separated list of source sheet names to include (optional).",
    )
    parser.add_argument(
        "--exclude",
        help="Comma-separated list of source sheet names to exclude (optional).",
    )
    parser.add_argument(
        "--source-prefix",
        default="",
        help="Only include sheets whose names start with this prefix (optional).",
    )
    parser.add_argument(
        "--no-source-column",
        action="store_true",
        help="Disable adding the Source Sheet column.",
    )
    parser.add_argument(
        "--source-column-name",
        default="Source Sheet",
        help="Name of source sheet column.",
    )
    parser.add_argument(
        "--no-clear-target",
        action="store_true",
        help="Do not clear target sheet before writing.",
    )
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else None

    rows_written, total_columns = union_orders_data(
        input_path=input_path,
        output_path=output_path,
        target_sheet_name=args.target_sheet,
        header_row=args.header_row,
        include_sheets=parse_csv_names(args.include),
        exclude_sheets=parse_csv_names(args.exclude),
        source_name_prefix=args.source_prefix,
        add_source_sheet_column=not args.no_source_column,
        source_sheet_column_name=args.source_column_name,
        clear_target_before_write=not args.no_clear_target,
    )

    destination = output_path if output_path else input_path
    print(
        f"Done. Wrote {rows_written} row(s) with {total_columns} column(s) "
        f"to sheet '{args.target_sheet}' in '{destination}'."
    )


if __name__ == "__main__":
    main()
