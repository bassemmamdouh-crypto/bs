#!/usr/bin/env python3
"""Build next-month item sales targets from area purchasing behavior.

For each area, the last N calendar months of sales are used to compute each
item's share of that area's total sales. Those shares (the mix) are applied
to the area's next-month total so SKU targets follow recent purchasing
behavior rather than a flat split.

Next-month area totals default to the average of the last N months. Override
them with ``--area-targets`` or an ``Area_Targets`` sheet in the input file.
"""

from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)

LOOKBACK_MONTHS = 3
DEFAULT_YEAR = 2026
_ARABIC_MONTH_RE = re.compile(r"(?:شهر|month)\s*[:\-]?\s*(\d{1,2})", re.IGNORECASE)
_YEAR_RE = re.compile(r"(20\d{2})")
_SKIP_MONTH_HEADERS = {
    "total",
    "grand_total",
    "sku",
    "item",
    "area",
    "الإجمالي",
    "اجمالي",
    "مبيعات_الجملة",
    "مبيعات_الجمله",
}
_SKIP_ITEM_NAMES = {"total", "grand total", "الإجمالي", "اجمالي"}

AREA_ALIASES = {
    "area",
    "area_name",
    "area_id",
    "polygon",
    "polygon_name",
    "city",
    "warehouse",
    "region",
    "zone",
    "location",
    "المنطقة",
    "منطقة",
}
ITEM_ALIASES = {
    "item",
    "item_name",
    "product",
    "product_name",
    "sku_name",
    "item_desc",
    "item_description",
    "الصنف",
    "الصنف_rout",
    "rout",
}
SKU_ALIASES = {"sku", "item_id", "product_id", "sku_id", "item_code"}
CONTAINER_ALIASES = {
    "container",
    "pack",
    "pack_size",
    "size",
    "العبوة",
    "العبوة_container",
}
SUPERVISOR_ALIASES = {
    "supervisor",
    "team",
    "اسم_المشرف",
    "اسم_المشرف_الفريق",
}
COST_CENTER_ALIASES = {
    "cost_center",
    "costcentre",
    "مركز_الكلفة",
    "مركز_كلفة",
}
MONTH_ALIASES = {
    "month",
    "year_month",
    "date",
    "sales_month",
    "period",
    "month_date",
    "order_month",
    "الشهر",
}
SALES_ALIASES = {
    "sales",
    "net_sales",
    "net_amount",
    "amount",
    "revenue",
    "gmv",
    "total_sales",
    "sales_value",
    "value",
    "net",
    "المبيعات",
}
QTY_ALIASES = {"qty", "quantity", "units", "sold_qty", "sales_qty"}
TARGET_ALIASES = {
    "next_month_target",
    "area_target",
    "target",
    "target_sales",
    "month_target",
    "next_month_sales_target",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6DCE4")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


class PlannerError(ValueError):
    """Raised when the input workbook cannot be interpreted."""


def normalize_column(name) -> str:
    text = str(name).strip().lower()
    text = re.sub(r"[^\w]+", "_", text, flags=re.UNICODE)
    return text.strip("_")


def _match_alias(columns: Iterable[str], aliases: set[str]) -> str | None:
    normalized = {normalize_column(col): col for col in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None


def parse_month_label(value, default_year: int | None = None) -> pd.Period | None:
    """Parse a cell or header into a monthly Period, or None if it is not a month."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, pd.Period):
        return value.asfreq("M")
    if isinstance(value, pd.Timestamp):
        return value.to_period("M")
    if hasattr(value, "year") and hasattr(value, "month") and not isinstance(value, str):
        try:
            return pd.Timestamp(value).to_period("M")
        except Exception:
            return None

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return None
    if normalize_column(text) in _SKIP_MONTH_HEADERS:
        return None

    year = default_year or DEFAULT_YEAR
    year_match = _YEAR_RE.search(text)
    if year_match:
        year = int(year_match.group(1))

    arabic = _ARABIC_MONTH_RE.search(text)
    if arabic:
        month_num = int(arabic.group(1))
        if 1 <= month_num <= 12:
            return pd.Period(year=year, month=month_num, freq="M")

    for fmt in ("%Y-%m", "%Y/%m", "%Y-%m-%d", "%Y/%m/%d", "%b %Y", "%B %Y", "%m/%Y", "%d/%m/%Y"):
        try:
            return pd.to_datetime(text, format=fmt).to_period("M")
        except (ValueError, TypeError):
            continue

    try:
        ts = pd.to_datetime(text, errors="raise")
        return ts.to_period("M")
    except (ValueError, TypeError, OverflowError):
        return None


def infer_year_from_workbook(path: str | Path) -> int:
    """Pick a calendar year from sheet titles / header cells, else DEFAULT_YEAR."""
    xls = pd.ExcelFile(path, engine="openpyxl")
    for name in xls.sheet_names:
        found = _YEAR_RE.search(str(name))
        if found:
            return int(found.group(1))
        peek = pd.read_excel(path, sheet_name=name, header=None, nrows=4, engine="openpyxl")
        for val in peek.astype(str).to_numpy().ravel():
            found = _YEAR_RE.search(val)
            if found:
                return int(found.group(1))
    return DEFAULT_YEAR


def _open_excel(path: str | Path) -> pd.ExcelFile:
    return pd.ExcelFile(path, engine="openpyxl")


def last_n_months(latest: pd.Period, n: int = LOOKBACK_MONTHS) -> list[pd.Period]:
    return [latest - (n - 1 - i) for i in range(n)]


def next_month(period: pd.Period) -> pd.Period:
    return period + 1


def _to_numeric(series: pd.Series) -> pd.Series:
    cleaned = series.astype(str).str.replace(",", "", regex=False).str.replace(" ", "", regex=False)
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)


def _detect_wide_month_columns(columns: Iterable, default_year: int | None = None) -> dict:
    months = {}
    for col in columns:
        period = parse_month_label(col, default_year=default_year)
        if period is not None:
            months[col] = period
    return months


def load_purchasing_table(
    path: str | Path,
    sheet: str | None = None,
    default_year: int | None = None,
) -> pd.DataFrame:
    """Load the first usable sheet (or a named sheet) as a DataFrame."""
    path = Path(path)
    if not path.exists():
        raise PlannerError(f"Input file not found: {path}")

    year = default_year or infer_year_from_workbook(path)
    workbook = _open_excel(path)
    if sheet:
        sheets = [sheet]
    else:
        preferred = []
        others = []
        for name in workbook.sheet_names:
            if name in {"Area_Targets", "Instructions", "معلومات"}:
                continue
            header = pd.read_excel(path, sheet_name=name, nrows=0, engine="openpyxl")
            cols = [normalize_column(c) for c in header.columns]
            if any(col in AREA_ALIASES or col == "المنطقة" for col in cols):
                preferred.append(name)
            else:
                others.append(name)
        sheets = preferred + others
    last_error = None
    for name in sheets:
        raw = pd.read_excel(path, sheet_name=name, engine="openpyxl")
        if raw.empty or raw.dropna(how="all").empty:
            continue
        try:
            return normalize_purchasing_frame(raw, default_year=year)
        except PlannerError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error
    raise PlannerError("No sheet with area / item / monthly sales was found.")


def normalize_purchasing_frame(
    raw: pd.DataFrame,
    default_year: int | None = None,
) -> pd.DataFrame:
    """Turn a long or wide purchasing sheet into area, item, sku, month, sales, qty."""
    if raw.empty:
        raise PlannerError("Purchasing sheet is empty.")

    raw = raw.copy()
    raw.columns = [str(c).strip() if not isinstance(c, pd.Timestamp) else c for c in raw.columns]
    raw = raw.dropna(how="all")

    area_col = _match_alias(raw.columns, AREA_ALIASES)
    item_col = _match_alias(raw.columns, ITEM_ALIASES)
    sku_col = _match_alias(raw.columns, SKU_ALIASES)
    container_col = _match_alias(raw.columns, CONTAINER_ALIASES)
    supervisor_col = _match_alias(raw.columns, SUPERVISOR_ALIASES)
    cost_center_col = _match_alias(raw.columns, COST_CENTER_ALIASES)
    month_col = _match_alias(raw.columns, MONTH_ALIASES)
    sales_col = _match_alias(raw.columns, SALES_ALIASES)
    qty_col = _match_alias(raw.columns, QTY_ALIASES)
    month_cols = _detect_wide_month_columns(raw.columns, default_year=default_year)

    if area_col is None:
        raise PlannerError(
            "Could not find an area column. Expected one of: "
            + ", ".join(sorted(AREA_ALIASES))
        )

    if item_col is None and sku_col is None:
        raise PlannerError(
            "Could not find an item or SKU column. Expected one of: "
            + ", ".join(sorted(ITEM_ALIASES | SKU_ALIASES))
        )

    if item_col is None:
        item_col = sku_col

    extra_cols = []
    for col in (supervisor_col, cost_center_col, container_col):
        if col and col not in extra_cols:
            extra_cols.append(col)

    id_cols = [area_col]
    if sku_col and sku_col != item_col and sku_col not in id_cols:
        id_cols.append(sku_col)
    if item_col not in id_cols:
        id_cols.append(item_col)
    for col in extra_cols:
        if col not in id_cols:
            id_cols.append(col)

    if month_col and sales_col:
        frame = raw[id_cols + [month_col, sales_col] + ([qty_col] if qty_col else [])].copy()
        frame["month"] = frame[month_col].map(lambda v: parse_month_label(v, default_year=default_year))
        frame["sales"] = _to_numeric(frame[sales_col])
        frame["qty"] = _to_numeric(frame[qty_col]) if qty_col else frame["sales"]
    elif month_cols:
        value_vars = list(month_cols)
        extra = [qty_col] if qty_col else []
        frame = raw[id_cols + value_vars + extra].melt(
            id_vars=id_cols + extra,
            value_vars=value_vars,
            var_name="_month_header",
            value_name="sales",
        )
        frame["month"] = frame["_month_header"].map(lambda h: month_cols[h])
        frame["sales"] = _to_numeric(frame["sales"])
        frame["qty"] = _to_numeric(frame[qty_col]) if qty_col else frame["sales"]
        frame = frame.drop(columns=["_month_header"])
    else:
        raise PlannerError(
            "Could not find monthly sales. Use long format (month + sales columns) "
            "or wide format (one column per month)."
        )

    frame = frame.rename(columns={area_col: "area", item_col: "item"})
    if supervisor_col:
        frame = frame.rename(columns={supervisor_col: "supervisor"})
    if cost_center_col:
        frame = frame.rename(columns={cost_center_col: "cost_center"})
    if container_col:
        frame = frame.rename(columns={container_col: "container"})
    if sku_col and sku_col in frame.columns and sku_col not in {"area", "item", "container"}:
        frame = frame.rename(columns={sku_col: "sku"})
    elif "container" in frame.columns:
        pack = frame["container"].astype(str).str.strip().replace({"nan": "", "None": ""})
        frame["sku"] = np.where(pack.eq(""), frame["item"].astype(str), pack + " | " + frame["item"].astype(str))
    else:
        frame["sku"] = frame["item"]

    frame["area"] = frame["area"].astype(str).str.strip()
    frame["item"] = frame["item"].astype(str).str.strip()
    frame["sku"] = frame["sku"].astype(str).str.strip()
    frame = frame.dropna(subset=["month"])
    frame = frame[(frame["area"] != "") & (frame["area"].str.lower() != "nan")]
    frame = frame[(frame["item"] != "") & (frame["item"].str.lower() != "nan")]
    frame = frame[~frame["item"].str.lower().isin(_SKIP_ITEM_NAMES)]

    group_cols = ["area", "sku", "item", "month"]
    for col in ("supervisor", "cost_center", "container"):
        if col in frame.columns:
            frame[col] = frame[col].astype(str).str.strip()
            group_cols.append(col)

    grouped = frame.groupby(group_cols, dropna=False, as_index=False)[["sales", "qty"]].sum()
    if grouped.empty:
        raise PlannerError("No usable area / item / month rows after cleaning.")
    return grouped


def load_area_targets(path: str | Path | None, sheet: str = "Area_Targets") -> pd.DataFrame:
    """Optional area -> next-month total sales overrides."""
    if path is None:
        return pd.DataFrame(columns=["area", "next_month_target"])
    path = Path(path)
    if not path.exists():
        raise PlannerError(f"Area-targets file not found: {path}")

    xls = pd.ExcelFile(path)
    sheet_name = sheet if sheet in xls.sheet_names else xls.sheet_names[0]
    raw = pd.read_excel(path, sheet_name=sheet_name)
    if raw.empty:
        return pd.DataFrame(columns=["area", "next_month_target"])

    area_col = _match_alias(raw.columns, AREA_ALIASES)
    target_col = _match_alias(raw.columns, TARGET_ALIASES)
    if area_col is None or target_col is None:
        raise PlannerError(
            "Area targets sheet needs an area column and a next-month target column."
        )
    out = pd.DataFrame(
        {
            "area": raw[area_col].astype(str).str.strip(),
            "next_month_target": _to_numeric(raw[target_col]),
        }
    )
    return out.dropna(subset=["area"]).groupby("area", as_index=False)["next_month_target"].sum()


def allocate_to_cost_centers(
    detail: pd.DataFrame,
    item_targets: pd.DataFrame,
    window: list[pd.Period],
) -> pd.DataFrame:
    """Split each area/item target across cost centers by their share of that item."""
    if "cost_center" not in detail.columns or detail.empty or item_targets.empty:
        return pd.DataFrame()

    keys = [c for c in ["area", "supervisor", "cost_center", "container", "sku", "item"] if c in detail.columns]
    grouped = (
        detail[detail["month"].isin(window)]
        .groupby(keys, as_index=False)
        .agg(cc_item_sales_last_n=("sales", "sum"))
    )
    grouped = grouped[grouped["cc_item_sales_last_n"] > 0]
    if grouped.empty:
        return pd.DataFrame()

    area_item = (
        grouped.groupby(["area", "sku", "item"], as_index=False)["cc_item_sales_last_n"]
        .sum()
        .rename(columns={"cc_item_sales_last_n": "area_item_sales_last_n"})
    )
    grouped = grouped.merge(area_item, on=["area", "sku", "item"], how="left")
    grouped["cost_center_share_of_item"] = (
        grouped["cc_item_sales_last_n"] / grouped["area_item_sales_last_n"].replace(0, np.nan)
    ).fillna(0.0)

    out = grouped.merge(
        item_targets[
            [
                "area",
                "sku",
                "item",
                "weighted_contribution",
                "next_month_target",
                "next_month_sales_target",
                "target_month",
            ]
        ].rename(columns={"next_month_sales_target": "area_item_target"}),
        on=["area", "sku", "item"],
        how="inner",
    )
    out["next_month_sales_target"] = out["area_item_target"] * out["cost_center_share_of_item"]
    out = out.sort_values(["area", "cost_center", "sku"]).reset_index(drop=True)
    return out


def build_mix_and_targets(
    purchasing: pd.DataFrame,
    area_targets: pd.DataFrame | None = None,
    lookback_months: int = LOOKBACK_MONTHS,
    latest_month: pd.Period | None = None,
) -> dict[str, pd.DataFrame]:
    """Compute last-N-month mix and next-month item targets.

    Mix used for targeting is the sales-weighted contribution:

        item_sales over last N months / area_sales over last N months

    Monthly shares and their simple average are kept as diagnostics so a
    volatile month is visible when setting the target.
    """
    if lookback_months < 1:
        raise PlannerError("lookback_months must be at least 1.")

    data = purchasing.copy()
    if latest_month is None:
        latest_month = data["month"].max()
    window = last_n_months(latest_month, lookback_months)
    target_month = next_month(latest_month)
    window_detail = data[data["month"].isin(window)].copy()
    if window_detail.empty:
        raise PlannerError(
            f"No sales rows in the last {lookback_months} months ending {latest_month}."
        )

    present_months = sorted(window_detail["month"].unique())
    n_present = max(len(present_months), 1)
    window_data = (
        window_detail.groupby(["area", "sku", "item", "month"], as_index=False)[["sales", "qty"]]
        .sum()
    )

    month_totals = (
        window_data.groupby(["area", "month"], as_index=False)["sales"]
        .sum()
        .rename(columns={"sales": "area_month_sales"})
    )
    monthly = window_data.merge(month_totals, on=["area", "month"], how="left")
    monthly["monthly_contribution"] = (
        monthly["sales"] / monthly["area_month_sales"].replace(0, np.nan)
    ).fillna(0.0)

    month_labels = [str(m) for m in window]
    contrib_wide = (
        monthly.pivot_table(
            index=["area", "sku", "item"],
            columns="month",
            values="monthly_contribution",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(columns=window, fill_value=0.0)
    )
    contrib_wide.columns = [f"contribution_{m}" for m in contrib_wide.columns]
    contrib_wide = contrib_wide.reset_index()

    sales_wide = (
        monthly.pivot_table(
            index=["area", "sku", "item"],
            columns="month",
            values="sales",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(columns=window, fill_value=0.0)
    )
    sales_wide.columns = [f"sales_{m}" for m in sales_wide.columns]
    sales_wide = sales_wide.reset_index()

    qty_wide = (
        monthly.pivot_table(
            index=["area", "sku", "item"],
            columns="month",
            values="qty",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(columns=window, fill_value=0.0)
    )
    qty_wide.columns = [f"qty_{m}" for m in qty_wide.columns]
    qty_wide = qty_wide.reset_index()

    item_3m = (
        window_data.groupby(["area", "sku", "item"], as_index=False)
        .agg(sales_last_n=("sales", "sum"), qty_last_n=("qty", "sum"))
    )
    area_3m = (
        window_data.groupby("area", as_index=False)
        .agg(area_sales_last_n=("sales", "sum"), area_qty_last_n=("qty", "sum"))
    )
    area_month_count = (
        month_totals[month_totals["area_month_sales"] > 0]
        .groupby("area")["month"]
        .nunique()
        .rename("months_with_sales")
        .reset_index()
    )

    items = item_3m.merge(area_3m, on="area", how="left")
    items = items.merge(area_month_count, on="area", how="left")
    items["months_with_sales"] = items["months_with_sales"].fillna(0).astype(int)
    items["weighted_contribution"] = (
        items["sales_last_n"] / items["area_sales_last_n"].replace(0, np.nan)
    ).fillna(0.0)
    items["avg_monthly_sales"] = items["sales_last_n"] / n_present
    items["avg_monthly_qty"] = items["qty_last_n"] / n_present
    items["unit_price"] = (
        items["sales_last_n"] / items["qty_last_n"].replace(0, np.nan)
    ).fillna(0.0)

    items = items.merge(contrib_wide, on=["area", "sku", "item"], how="left")
    items = items.merge(sales_wide, on=["area", "sku", "item"], how="left")
    items = items.merge(qty_wide, on=["area", "sku", "item"], how="left")
    contrib_cols = [f"contribution_{m}" for m in window]
    items[contrib_cols] = items[contrib_cols].fillna(0.0)
    contrib_present = [f"contribution_{m}" for m in present_months]
    items["simple_avg_contribution"] = items[contrib_present].mean(axis=1)

    area_avg = (
        window_data.groupby("area")["sales"].sum().div(n_present).rename("default_next_month_target")
    )
    area_summary = area_3m.merge(area_avg, on="area", how="left")
    area_summary = area_summary.merge(area_month_count, on="area", how="left")
    area_summary["months_with_sales"] = area_summary["months_with_sales"].fillna(0).astype(int)
    area_summary["lookback_start"] = str(present_months[0])
    area_summary["lookback_end"] = str(present_months[-1])
    area_summary["target_month"] = str(target_month)
    area_summary["avg_monthly_sales"] = area_summary["area_sales_last_n"] / n_present
    area_summary["months_in_file"] = n_present
    area_summary["lookback_requested"] = lookback_months

    overrides = area_targets if area_targets is not None else pd.DataFrame(columns=["area", "next_month_target"])
    if not overrides.empty:
        area_summary = area_summary.merge(overrides, on="area", how="left")
    else:
        area_summary["next_month_target"] = pd.NA
    area_summary["target_source"] = area_summary["next_month_target"].apply(
        lambda v: "override" if pd.notna(v) else f"{n_present}-month average"
    )
    area_summary["next_month_target"] = area_summary["next_month_target"].fillna(
        area_summary["default_next_month_target"]
    )

    items = items.merge(
        area_summary[["area", "next_month_target", "target_month", "lookback_start", "lookback_end"]],
        on="area",
        how="left",
    )
    items["next_month_sales_target"] = items["next_month_target"] * items["weighted_contribution"]
    items["next_month_qty_target"] = (
        items["next_month_sales_target"] / items["unit_price"].replace(0, np.nan)
    ).fillna(0.0)
    items = items[items["sales_last_n"] > 0].copy()
    items["rank_in_area"] = (
        items.groupby("area")["weighted_contribution"].rank(method="first", ascending=False).astype(int)
    )
    items = items.sort_values(["area", "rank_in_area", "item"]).reset_index(drop=True)

    mix_long = monthly[monthly["sales"] > 0].copy()
    mix_long["target_month"] = str(target_month)

    cost_center_targets = allocate_to_cost_centers(window_detail, items, window)

    return {
        "area_summary": area_summary.sort_values("area").reset_index(drop=True),
        "item_targets": items,
        "monthly_mix": mix_long.sort_values(["area", "month", "item"]).reset_index(drop=True),
        "cost_center_targets": cost_center_targets,
        "window_months": month_labels,
        "present_months": [str(m) for m in present_months],
        "target_month": str(target_month),
        "latest_month": str(latest_month),
        "lookback_months": lookback_months,
        "months_in_file": n_present,
    }


def _style_header(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 30
    ws.auto_filter.ref = f"A{row}:{get_column_letter(col_count)}{row}"
    ws.freeze_panes = f"A{row + 1}"


def _autosize(ws: Worksheet, max_width: int = 36) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(12, length + 2)


def _write_dataframe(ws: Worksheet, df: pd.DataFrame, start_row: int = 1) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_excel_value(value))
            cell.border = THIN
            if r_idx > start_row and (r_idx - start_row) % 2 == 0:
                cell.fill = ALT_FILL
    if not df.empty:
        _style_header(ws, start_row, len(df.columns))
        _apply_number_formats(ws, df, start_row)
    _autosize(ws)


def _excel_value(value):
    if isinstance(value, pd.Period):
        return str(value)
    if pd.isna(value):
        return None
    if hasattr(value, "item") and not isinstance(value, (bytes, str)):
        try:
            return value.item()
        except (ValueError, AttributeError):
            return value
    return value


def _apply_number_formats(ws: Worksheet, df: pd.DataFrame, header_row: int) -> None:
    pct_tokens = ("contribution",)
    money_tokens = ("sales", "target", "price", "amount")
    qty_tokens = ("qty", "rank", "months")
    for col_idx, name in enumerate(df.columns, start=1):
        key = str(name).lower()
        if any(tok in key for tok in pct_tokens):
            fmt = "0.00%"
        elif any(tok in key for tok in money_tokens):
            fmt = "#,##0.00"
        elif any(tok in key for tok in qty_tokens):
            fmt = "#,##0.00" if "qty" in key else "0"
        else:
            continue
        for row in range(header_row + 1, header_row + 1 + len(df)):
            ws.cell(row=row, column=col_idx).number_format = fmt


def _instructions_text(result: dict) -> list[str]:
    present = result.get("present_months") or result["window_months"]
    missing_note = ""
    if result.get("months_in_file", len(present)) < result["lookback_months"]:
        missing_note = (
            f"Only {result['months_in_file']} month(s) are in the file: {', '.join(present)}. "
            "Contribution and the default area total use those months. Empty months are not treated as zero sales."
        )
    return [
        "Area item target planner",
        "",
        f"Target month: {result['target_month']}",
        f"Requested lookback: last {result['lookback_months']} months ({result['window_months'][0]} to {result['window_months'][-1]})",
        f"Months used: {', '.join(present)}",
        missing_note,
        "",
        "How targets are set",
        "1. For each area, take sales in the months that exist in the lookback window.",
        "2. Item contribution = item sales in those months / area total sales in those months.",
        "3. Next-month item target = area next-month total × item contribution.",
        "",
        "Area next-month totals default to average monthly sales over the months actually present.",
        "Type a different total in Area_Targets.next_month_target and rerun with --area-targets to override.",
        "Cost_Center_Targets splits each area/item target by that cost center's share of the item.",
        "",
        "Sheets",
        "Area_Targets — one row per area; edit next_month_target then rerun.",
        "Item_Targets — SKU mix and suggested next-month targets for every area.",
        "Cost_Center_Targets — same item targets split to cost centers / distributors.",
        "Monthly_Mix — each present month's contribution, for audit.",
        "One sheet per area.",
    ]


def write_output_workbook(result: dict[str, pd.DataFrame], path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Instructions"
    for i, line in enumerate(_instructions_text(result), start=1):
        ws_info.cell(row=i, column=1, value=line)
        if i == 1:
            ws_info.cell(row=i, column=1).font = TITLE_FONT
        elif line in {"How targets are set", "Sheets"}:
            ws_info.cell(row=i, column=1).fill = SECTION_FILL
            ws_info.cell(row=i, column=1).font = Font(bold=True)
    ws_info.column_dimensions["A"].width = 110

    area_cols = [
        "area",
        "target_month",
        "lookback_start",
        "lookback_end",
        "lookback_requested",
        "months_in_file",
        "months_with_sales",
        "area_sales_last_n",
        "avg_monthly_sales",
        "default_next_month_target",
        "next_month_target",
        "target_source",
    ]
    item_cols = [
        "area",
        "sku",
        "item",
        "rank_in_area",
        "target_month",
        "lookback_start",
        "lookback_end",
        "weighted_contribution",
        "simple_avg_contribution",
        *[f"contribution_{m}" for m in result["window_months"]],
        "sales_last_n",
        "avg_monthly_sales",
        "next_month_target",
        "next_month_sales_target",
        "qty_last_n",
        "avg_monthly_qty",
        "unit_price",
        "next_month_qty_target",
    ]
    mix_cols = [
        "area",
        "sku",
        "item",
        "month",
        "sales",
        "qty",
        "area_month_sales",
        "monthly_contribution",
        "target_month",
    ]

    ws_area = wb.create_sheet("Area_Targets")
    area_df = result["area_summary"].reindex(columns=area_cols)
    _write_dataframe(ws_area, area_df)

    ws_items = wb.create_sheet("Item_Targets")
    item_df = result["item_targets"].reindex(columns=[c for c in item_cols if c in result["item_targets"].columns])
    _write_dataframe(ws_items, item_df)

    cc = result.get("cost_center_targets")
    if cc is not None and not cc.empty:
        ws_cc = wb.create_sheet("Cost_Center_Targets")
        cc_cols = [
            "area",
            "supervisor",
            "cost_center",
            "container",
            "sku",
            "item",
            "target_month",
            "weighted_contribution",
            "cost_center_share_of_item",
            "cc_item_sales_last_n",
            "area_item_target",
            "next_month_sales_target",
            "next_month_target",
        ]
        _write_dataframe(ws_cc, cc.reindex(columns=[c for c in cc_cols if c in cc.columns]))

    ws_mix = wb.create_sheet("Monthly_Mix")
    mix_df = result["monthly_mix"].copy()
    mix_df["month"] = mix_df["month"].astype(str)
    mix_df = mix_df.reindex(columns=mix_cols)
    _write_dataframe(ws_mix, mix_df)

    # One sheet per area so planners can work area by area.
    for area, group in result["item_targets"].groupby("area", sort=True):
        title = re.sub(r"[\\/*?:\[\]]", "_", str(area))[:31] or "Area"
        original = title
        suffix = 1
        while title in wb.sheetnames:
            suffix += 1
            title = f"{original[:28]}_{suffix}"
        ws = wb.create_sheet(title)
        subset = group.reindex(columns=[c for c in item_cols if c in group.columns])
        _write_dataframe(ws, subset)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def plan_from_excel(
    input_path: str | Path,
    output_path: str | Path,
    area_targets_path: str | Path | None = None,
    lookback_months: int = LOOKBACK_MONTHS,
    sheet: str | None = None,
    default_year: int | None = None,
) -> dict:
    purchasing = load_purchasing_table(input_path, sheet=sheet, default_year=default_year)
    overrides = None
    if area_targets_path:
        overrides = load_area_targets(area_targets_path)
    else:
        try:
            xls = _open_excel(input_path)
            if "Area_Targets" in xls.sheet_names:
                overrides = load_area_targets(input_path, sheet="Area_Targets")
        except Exception:
            overrides = None
    result = build_mix_and_targets(
        purchasing,
        area_targets=overrides,
        lookback_months=lookback_months,
    )
    write_output_workbook(result, output_path)
    result["input_rows"] = len(purchasing)
    result["output_path"] = str(output_path)
    return result


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Set next-month item targets from last-3-month area sales mix."
    )
    parser.add_argument("input", help="Excel file with purchasing behavior (long or wide).")
    parser.add_argument(
        "-o",
        "--output",
        default="area_item_targets.xlsx",
        help="Output workbook path (default: area_item_targets.xlsx).",
    )
    parser.add_argument(
        "--area-targets",
        default=None,
        help="Optional Excel with area, next_month_target to override default totals.",
    )
    parser.add_argument(
        "--months",
        type=int,
        default=LOOKBACK_MONTHS,
        help="How many trailing months to use for mix (default: 3).",
    )
    parser.add_argument("--sheet", default=None, help="Purchasing sheet name if the file has several.")
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Calendar year for headers like شهر 6 (default: detect from the file, else 2026).",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO, format="%(message)s")
    result = plan_from_excel(
        args.input,
        args.output,
        area_targets_path=args.area_targets,
        lookback_months=args.months,
        sheet=args.sheet,
        default_year=args.year,
    )
    items = result["item_targets"]
    areas = result["area_summary"]
    logger.info(
        "Wrote %s | target month %s | %s areas | %s items | lookback %s–%s",
        result["output_path"],
        result["target_month"],
        areas["area"].nunique(),
        len(items),
        result["window_months"][0],
        result["window_months"][-1],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
