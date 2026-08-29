#!/usr/bin/env python3
"""Write a sample purchasing-behavior workbook (wide and long layouts)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def create_sample(path: str | Path = "sample_purchasing_behavior.xlsx") -> Path:
    path = Path(path)
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    # Wide layout — common export: area, item, one column per month.
    wide = wb.active
    wide.title = "Purchasing_Wide"
    headers = ["area", "sku", "item", "2026-05", "2026-06", "2026-07"]
    for col, name in enumerate(headers, start=1):
        cell = wide.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        ("Baghdad", "P330", "Pepsi 330ml", 120000, 130000, 125000),
        ("Baghdad", "P1500", "Pepsi 1.5L", 80000, 70000, 90000),
        ("Baghdad", "L25", "Lays Classic 25g", 40000, 45000, 50000),
        ("Baghdad", "L40", "Lays Salt 40g", 20000, 15000, 10000),
        ("Basra", "P330", "Pepsi 330ml", 50000, 55000, 60000),
        ("Basra", "P1500", "Pepsi 1.5L", 30000, 28000, 32000),
        ("Basra", "L25", "Lays Classic 25g", 18000, 20000, 22000),
        ("Erbil", "P330", "Pepsi 330ml", 40000, 42000, 41000),
        ("Erbil", "L25", "Lays Classic 25g", 25000, 24000, 26000),
        ("Erbil", "C50", "Cheetos 50g", 10000, 8000, 9000),
    ]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            wide.cell(row=r_idx, column=c_idx, value=value)

    # Long layout with qty so quantity targets can be derived.
    long = wb.create_sheet("Purchasing_Long")
    long_headers = ["area", "sku", "item", "month", "sales", "qty"]
    for col, name in enumerate(long_headers, start=1):
        cell = long.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    prices = {"P330": 500, "P1500": 1000, "L25": 250, "L40": 400, "C50": 300}
    r = 2
    for area, sku, item, m5, m6, m7 in rows:
        price = prices[sku]
        for month, sales in (("2026-05", m5), ("2026-06", m6), ("2026-07", m7)):
            long.cell(row=r, column=1, value=area)
            long.cell(row=r, column=2, value=sku)
            long.cell(row=r, column=3, value=item)
            long.cell(row=r, column=4, value=month)
            long.cell(row=r, column=5, value=sales)
            long.cell(row=r, column=6, value=round(sales / price, 2))
            r += 1

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Sample purchasing behavior"
    notes["A1"].font = Font(bold=True, size=14)
    notes["A3"] = (
        "Last 3 months are May–July 2026. The planner sets August 2026 item targets "
        "from each area's sales mix over those months."
    )
    notes["A5"] = "Use either Purchasing_Wide or Purchasing_Long as the input sheet."
    notes.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = create_sample()
    print(f"Wrote {out}")
